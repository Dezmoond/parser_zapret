"""
Парсер 5 - Парсер реестра иностранных агентов Минюста
"""
import logging
import time
import os
from typing import List, Dict, Any
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import openpyxl
import csv
from .base_parser import BaseParser


class Parser5(BaseParser):
    """Парсер 5 - парсит XLSX файл реестра иностранных агентов с сайта Минюста"""
    
    def __init__(self, config: Dict[str, Any]):
        """Инициализация парсера с настройками Selenium"""
        super().__init__(config)
        self.use_selenium = config.get('use_selenium', True)
        self.selenium_timeout = config.get('selenium_timeout', 30)
        self.download_dir = config.get('download_dir', 'downloads')
        
    def _init_selenium_driver(self):
        """Инициализация Selenium WebDriver с настройками для скачивания файлов"""
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument(f'user-agent={self.user_agent}')
        chrome_options.add_argument('--disable-blink-features=AutomationControlled')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        
        # Настройки для скачивания файлов
        prefs = {
            "download.default_directory": str(Path(self.download_dir).absolute()),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        try:
            try:
                from selenium.webdriver.chrome.service import Service
                from webdriver_manager.chrome import ChromeDriverManager
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
                self.logger.info("Использован webdriver-manager для ChromeDriver")
            except ImportError:
                driver = webdriver.Chrome(options=chrome_options)
                self.logger.info("Использован стандартный ChromeDriver")
            
            driver.set_page_load_timeout(self.selenium_timeout)
            return driver
        except Exception as e:
            self.logger.error(f"Ошибка инициализации Selenium: {e}")
            raise
    
    def _download_xlsx(self, url: str) -> str:
        """
        Скачивание XLSX файла с сайта с переименованием
        
        Args:
            url: URL страницы
            
        Returns:
            Путь к скачанному и переименованному файлу
        """
        driver = None
        downloaded_file = None
        
        try:
            # Создаем директорию для загрузок
            download_path = Path(self.download_dir)
            download_path.mkdir(parents=True, exist_ok=True)
            
            driver = self._init_selenium_driver()
            self.logger.info(f"Загрузка страницы: {url}")
            driver.get(url)
            
            # Ждем загрузки страницы
            time.sleep(3)
            
            # Ищем кнопку для скачивания
            try:
                download_button = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#registry_download_xls'))
                )
                self.logger.info("Кнопка скачивания найдена")
                
                # Получаем список файлов до клика
                files_before = set(download_path.glob('*.xlsx'))
                
                # Кликаем на кнопку
                driver.execute_script("arguments[0].click();", download_button)
                self.logger.info("Кнопка скачивания нажата")
                
                # Ждем скачивания файла (максимум 60 секунд)
                max_wait = 60
                waited = 0
                while waited < max_wait:
                    time.sleep(1)
                    waited += 1
                    files_after = set(download_path.glob('*.xlsx'))
                    new_files = files_after - files_before
                    if new_files:
                        downloaded_file = str(list(new_files)[0])
                        self.logger.info(f"Файл скачан: {downloaded_file}")
                        break
                
                if not downloaded_file:
                    # Пробуем найти последний измененный файл
                    xlsx_files = list(download_path.glob('*.xlsx'))
                    if xlsx_files:
                        downloaded_file = str(max(xlsx_files, key=os.path.getmtime))
                        self.logger.info(f"Использован последний файл: {downloaded_file}")
                
                # Переименовываем файл, чтобы избежать конфликта с parser2
                if downloaded_file:
                    original_path = Path(downloaded_file)
                    new_name = f"parser5_{original_path.name}"
                    new_path = original_path.parent / new_name
                    
                    # Если файл с таким именем уже существует, удаляем его
                    if new_path.exists():
                        os.remove(new_path)
                    
                    os.rename(original_path, new_path)
                    downloaded_file = str(new_path)
                    self.logger.info(f"Файл переименован в: {downloaded_file}")
                
            except (TimeoutException, NoSuchElementException) as e:
                self.logger.error(f"Не удалось найти или нажать кнопку скачивания: {e}")
                raise
            
        except Exception as e:
            self.logger.error(f"Ошибка при скачивании файла: {e}")
            raise
        finally:
            if driver:
                driver.quit()
        
        if not downloaded_file:
            raise FileNotFoundError("XLSX файл не был скачан")
        
        return downloaded_file
    
    def _parse_xlsx(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Парсинг XLSX файла, пропуская первые 2 строки
        
        Args:
            file_path: Путь к XLSX файлу
            
        Returns:
            Список словарей с данными
        """
        data = []
        
        try:
            self.logger.info(f"Чтение XLSX файла: {file_path}")
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Пропускаем первые 2 строки
            # Строка 1 (индекс 1) - пропускаем
            # Строка 2 (индекс 2) - пропускаем
            # Строка 3 (индекс 3) - заголовки таблицы
            # Строка 4 (индекс 4) - начало данных
            
            # Определяем максимальное количество столбцов
            max_col = sheet.max_column
            
            # Читаем заголовки из 3-й строки
            header_row = sheet[3]  # 3-я строка (индекс 3)
            headers = []
            
            for i, cell in enumerate(header_row):
                if cell.value:
                    header = str(cell.value).strip()
                else:
                    header = f'Column_{i+1}'
                headers.append(header)
            
            # Если заголовки не найдены, создаем по умолчанию
            if not headers or all(not h or h.startswith('Column_') for h in headers):
                headers = [f'Column_{i+1}' for i in range(max_col)]
            
            # Данные начинаются с 4-й строки (индекс 4)
            start_row = 4
            
            self.logger.info(f"Заголовки: {headers}")
            self.logger.info(f"Начало данных со строки: {start_row}")
            
            # Читаем данные начиная с start_row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=False), start=start_row):
                # Пропускаем пустые строки
                if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                    continue
                
                # Создаем словарь для строки
                row_data = {}
                for i, cell in enumerate(row):
                    header = headers[i] if i < len(headers) else f'Column_{i+1}'
                    value = cell.value
                    if value is not None:
                        # Преобразуем значение в строку и очищаем
                        value = str(value).strip()
                    else:
                        value = ''
                    row_data[header] = value
                
                # Добавляем только если есть хотя бы одно непустое значение
                if any(row_data.values()):
                    data.append(row_data)
            
            self.logger.info(f"Прочитано {len(data)} строк данных из XLSX")
            
        except Exception as e:
            self.logger.error(f"Ошибка парсинга XLSX файла: {e}")
            raise
        
        return data
    
    def parse(self, response) -> List[Dict[str, Any]]:
        """
        Этот метод не используется для этого парсера, так как мы переопределяем run()
        """
        return []
    
    def run(self) -> None:
        """
        Переопределенный метод run для скачивания и парсинга XLSX файла
        """
        if not self.enabled:
            self.logger.info(f"Парсер {self.name} отключен в конфигурации")
            return
        
        self.logger.info(f"Запуск парсера {self.name}")
        
        if not self.urls:
            self.logger.warning("Нет URL для парсинга")
            return
        
        url = self.urls[0]
        
        try:
            # Скачиваем XLSX файл
            xlsx_file = self._download_xlsx(url)
            
            # Парсим XLSX файл
            data = self._parse_xlsx(xlsx_file)
            
            # Проверяем наличие данных
            if not data or len(data) == 0:
                self.logger.warning(f"⚠️  ВНИМАНИЕ: Парсер {self.name} не собрал данных из файла {xlsx_file}")
                
                from pathlib import Path
                if Path(self.output_file).exists():
                    self.logger.warning(f"Существующий файл {self.output_file} НЕ будет перезаписан пустыми данными")
                else:
                    self.logger.warning(f"Файл {self.output_file} не будет создан")
                return
            
            # Сохраняем в CSV
            self.save_to_csv(data, self.output_file)
            self.logger.info(f"✓ Сохранено {len(data)} записей в {self.output_file}")
            
            # Удаляем скачанный XLSX файл (опционально)
            try:
                if os.path.exists(xlsx_file):
                    os.remove(xlsx_file)
                    self.logger.info(f"Временный файл {xlsx_file} удален")
            except Exception as e:
                self.logger.warning(f"Не удалось удалить временный файл {xlsx_file}: {e}")
            
        except Exception as e:
            self.logger.error(f"Ошибка при работе парсера: {e}")
            from pathlib import Path
            if Path(self.output_file).exists():
                self.logger.warning(f"Существующий файл {self.output_file} НЕ будет перезаписан")
    
    def save_to_csv(self, data: List[Dict[str, Any]], output_file: str = None) -> None:
        """
        Сохранение данных в CSV файл
        
        Args:
            data: Список словарей с данными
            output_file: Путь к выходному файлу (если None, используется self.output_file)
        """
        if not data:
            self.logger.warning("Нет данных для сохранения")
            return
        
        import csv
        
        if output_file is None:
            output_file = self.output_file
        
        # Создаем директорию если нужно
        from pathlib import Path
        output_path = Path(output_file)
        output_path = output_path.resolve()
        
        try:
            self.logger.info(f"Создание директории: {output_path.parent}")
            output_path.parent.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"Директория создана/проверена: {output_path.parent}")
        except Exception as e:
            self.logger.error(f"Ошибка при создании директории {output_path.parent}: {e}")
            raise
        
        # Получаем все уникальные ключи из всех записей
        fieldnames = set()
        for record in data:
            fieldnames.update(record.keys())
        
        # Номер всегда должен быть первым столбцом
        fieldnames_list = list(fieldnames)
        number_field = None
        
        # Ищем поле с номером
        for field in fieldnames_list:
            field_lower = field.lower().replace(' ', '').replace('\xa0', '')
            if ('№' in field or 'номер' in field_lower or 'п/п' in field_lower or 
                field_lower.startswith('№') or 'пп' in field_lower or 
                field == '№п/п' or field == '№ п/п' or field.replace(' ', '') == '№п/п'):
                number_field = field
                self.logger.info(f"Найдено поле с номером: '{field}'")
                break
        
        if number_field:
            fieldnames_list.remove(number_field)
            fieldnames = [number_field] + sorted(fieldnames_list)
        else:
            fieldnames = sorted(fieldnames_list)
        
        # Записываем данные в CSV
        with open(output_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        
        self.logger.info(f"Данные сохранены в {output_file}")

